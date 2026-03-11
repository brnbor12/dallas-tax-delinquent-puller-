#!/usr/bin/env python3
"""
Deal Calculator Workbook Generator
===================================
Generates a multi-sheet Excel underwriting workbook for Dallas
tax-delinquent wholesale deals using openpyxl.

Sheets produced:
  SETTINGS        – core assumptions (sell costs, hold costs, margins)
  DEAL_CALC       – main underwriting: ARV, repairs, MAO, offers
  COMPS           – comp-based ARV builder (up to 6 comps)
  REPAIRS         – quick tier + detailed line-item rehab estimator
  DISPO_SUMMARY   – one-page buyer snapshot
"""

import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
DARK_BLUE = "1F3864"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ACCENT_GREEN = "548235"
ACCENT_RED = "C00000"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
BLACK = "000000"

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=12)
SECTION_FONT = Font(name="Calibri", bold=True, color=DARK_BLUE, size=11)
LABEL_FONT = Font(name="Calibri", size=10)
INPUT_FONT = Font(name="Calibri", size=10, color=MED_BLUE)
RESULT_FONT = Font(name="Calibri", bold=True, size=11)

HEADER_FILL = PatternFill("solid", fgColor=DARK_BLUE)
SECTION_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow
RESULT_FILL = PatternFill("solid", fgColor="E2EFDA")  # light green
WARN_FILL = PatternFill("solid", fgColor="FCE4EC")  # light red

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CURRENCY_FMT = '"$"#,##0'
CURRENCY_FMT2 = '"$"#,##0.00'
PCT_FMT = "0.0%"
PCT_FMT0 = "0%"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _header_row(ws, row, text, cols=6):
    """Merge and style a dark-blue header row."""
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=cols
    )
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28


def _section_row(ws, row, text, cols=6):
    """Merge and style a light-blue section header."""
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=cols
    )
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22


def _label(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = LABEL_FONT
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = THIN_BORDER
    return cell


def _input(ws, row, col, value=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = INPUT_FONT
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt
    return cell


def _formula(ws, row, col, formula_str, fmt=None, bold=False):
    cell = ws.cell(row=row, column=col, value=formula_str)
    cell.font = Font(
        name="Calibri", size=10, bold=bold, color=ACCENT_GREEN if bold else BLACK
    )
    cell.fill = RESULT_FILL if bold else PatternFill()
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt
    return cell


def _result(ws, row, col, formula_str, fmt=None):
    return _formula(ws, row, col, formula_str, fmt=fmt, bold=True)


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def build_settings(wb):
    """SETTINGS sheet – all core assumptions in one place."""
    ws = wb.create_sheet("SETTINGS")
    _set_col_widths(ws, [32, 16, 6, 32, 16, 6])

    _header_row(ws, 1, "SETTINGS – Core Assumptions", cols=6)

    # --- Sell / disposition costs ---
    _section_row(ws, 3, "Sell-Side Costs (as % of ARV)", cols=6)
    labels_sell = [
        ("Realtor commission", 0.05),
        ("Seller closing costs", 0.02),
        ("Holding cost (monthly)", 0.01),
        ("Est. hold period (months)", 6),
    ]
    for i, (lbl, val) in enumerate(labels_sell):
        r = 4 + i
        _label(ws, r, 1, lbl)
        if "period" in lbl.lower():
            fmt = "#,##0"
        else:
            fmt = PCT_FMT
        _input(ws, r, 2, val, fmt=fmt)

    # --- Wholesale margins ---
    _section_row(ws, 9, "Wholesale / Offer Margins", cols=6)
    margin_labels = [
        ("Classic MAO % of ARV", 0.70),
        ("Tight deal MAO %", 0.65),
        ("Target buyer profit %", 0.15),
        ("Min assignment fee ($)", 10000),
        ("Opening offer discount %", 0.20),
        ("Target offer discount %", 0.10),
        ("Absolute max buffer %", 0.02),
    ]
    for i, (lbl, val) in enumerate(margin_labels):
        r = 10 + i
        _label(ws, r, 1, lbl)
        fmt = PCT_FMT if "%" in lbl else CURRENCY_FMT
        _input(ws, r, 2, val, fmt=fmt)

    # --- Quick rehab tiers ---
    _section_row(ws, 18, "Quick Rehab Cost-per-SqFt Tiers", cols=6)
    tiers = [
        ("Light cosmetic ($/sqft)", 15),
        ("Moderate rehab ($/sqft)", 30),
        ("Heavy rehab ($/sqft)", 50),
        ("Full gut / structural ($/sqft)", 75),
    ]
    for i, (lbl, val) in enumerate(tiers):
        r = 19 + i
        _label(ws, r, 1, lbl)
        _input(ws, r, 2, val, fmt=CURRENCY_FMT)

    # --- Buy box filters (for future v2) ---
    _section_row(ws, 24, "Dallas Buy-Box Filters (v2)", cols=6)
    filters = [
        ("Target zip codes", "75215, 75216, 75217, 75223, 75227"),
        ("Min ARV ($)", 80000),
        ("Max ARV ($)", 250000),
        ("Max seller ask ($)", 150000),
        ("Max rehab ($)", 60000),
    ]
    for i, (lbl, val) in enumerate(filters):
        r = 25 + i
        _label(ws, r, 1, lbl)
        fmt = CURRENCY_FMT if isinstance(val, (int, float)) else None
        _input(ws, r, 2, val, fmt=fmt)

    ws.sheet_properties.tabColor = DARK_BLUE
    return ws


def build_comps(wb):
    """COMPS sheet – comp-based ARV builder for up to 6 comps."""
    ws = wb.create_sheet("COMPS")
    _set_col_widths(ws, [24, 18, 14, 14, 14, 14, 14, 14])

    _header_row(ws, 1, "COMPS – Comparable Sales ARV Builder", cols=8)

    headers = [
        "Field", "Comp 1", "Comp 2", "Comp 3",
        "Comp 4", "Comp 5", "Comp 6",
    ]
    for i, h in enumerate(headers):
        c = ws.cell(row=3, column=i + 1, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=MED_BLUE)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")

    fields = [
        "Address",
        "Sale price",
        "Sq ft",
        "Price / sqft",
        "Bed / Bath",
        "Year built",
        "Days on market",
        "Condition adj (+/-)",
        "Adjusted price",
    ]
    for i, f in enumerate(fields):
        r = 4 + i
        _label(ws, r, 1, f)
        for comp_col in range(2, 8):
            if f == "Price / sqft":
                col_letter = get_column_letter(comp_col)
                _formula(
                    ws, r, comp_col,
                    f'=IF({col_letter}5=0,"",{col_letter}5/{col_letter}6)',
                    fmt=CURRENCY_FMT2,
                )
            elif f == "Adjusted price":
                col_letter = get_column_letter(comp_col)
                _formula(
                    ws, r, comp_col,
                    f"=IF({col_letter}5=0,0,{col_letter}5+{col_letter}11)",
                    fmt=CURRENCY_FMT,
                )
            else:
                fmt = CURRENCY_FMT if "price" in f.lower() or "adj" in f.lower() else None
                _input(ws, r, comp_col, fmt=fmt)

    # Summary row
    r_summary = 14
    _section_row(ws, r_summary, "ARV Summary", cols=8)

    _label(ws, r_summary + 1, 1, "Average adjusted price")
    _result(
        ws, r_summary + 1, 2,
        "=AVERAGEIF(B12:G12,\">0\")",
        fmt=CURRENCY_FMT,
    )

    _label(ws, r_summary + 2, 1, "Median adjusted price")
    _result(
        ws, r_summary + 2, 2,
        "=MEDIAN(IF(B12:G12>0,B12:G12))",
        fmt=CURRENCY_FMT,
    )

    _label(ws, r_summary + 3, 1, "Comp-based ARV (avg)")
    _result(
        ws, r_summary + 3, 2,
        "=B15",
        fmt=CURRENCY_FMT,
    )

    ws.sheet_properties.tabColor = MED_BLUE
    return ws


def build_repairs(wb):
    """REPAIRS sheet – quick tier estimate + detailed line-item rehab."""
    ws = wb.create_sheet("REPAIRS")
    _set_col_widths(ws, [30, 16, 16, 16, 16])

    _header_row(ws, 1, "REPAIRS – Rehab Cost Estimator", cols=5)

    # --- Quick estimate ---
    _section_row(ws, 3, "Quick Estimate (by condition tier)", cols=5)

    _label(ws, 4, 1, "Square footage")
    _input(ws, 4, 2, 1400, fmt="#,##0")

    _label(ws, 5, 1, "Condition tier")
    # Dropdown placeholder — user picks Light / Moderate / Heavy / Gut
    _input(ws, 5, 2, "Moderate")

    _label(ws, 6, 1, "Cost per sqft (from SETTINGS)")
    _formula(
        ws, 6, 2,
        '=IF(B5="Light",SETTINGS!B19,IF(B5="Moderate",SETTINGS!B20,'
        'IF(B5="Heavy",SETTINGS!B21,SETTINGS!B22)))',
        fmt=CURRENCY_FMT,
    )

    _label(ws, 7, 1, "Quick rehab total")
    _result(ws, 7, 2, "=B4*B6", fmt=CURRENCY_FMT)

    # --- Detailed line items ---
    _section_row(ws, 9, "Detailed Line-Item Rehab", cols=5)
    detail_headers = ["Category", "Description", "Qty", "Unit Cost", "Line Total"]
    for i, h in enumerate(detail_headers):
        c = ws.cell(row=10, column=i + 1, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=MED_BLUE)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")

    line_items = [
        ("Roof", "Tear-off + reshingle", 1, 7500),
        ("HVAC", "Replace full system", 1, 5500),
        ("Plumbing", "Repipe / repair", 1, 3500),
        ("Electrical", "Panel + rewire", 1, 4000),
        ("Foundation", "Pier / level", 1, 5000),
        ("Exterior paint", "Full exterior", 1, 3000),
        ("Interior paint", "Full interior", 1, 2500),
        ("Flooring", "LVP / tile install", 1400, 3),
        ("Kitchen", "Cabinets + counters + appliances", 1, 8000),
        ("Bathrooms", "Full bath reno x2", 2, 4000),
        ("Windows", "Replace all", 10, 350),
        ("Landscaping", "Clean-up + sod", 1, 1500),
        ("Dumpster / demo", "Debris removal", 1, 2000),
        ("Permits / misc", "City permits + contingency", 1, 2000),
    ]
    for i, (cat, desc, qty, unit_cost) in enumerate(line_items):
        r = 11 + i
        _label(ws, r, 1, cat)
        ws.cell(row=r, column=2, value=desc).font = LABEL_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        _input(ws, r, 3, qty, fmt="#,##0")
        _input(ws, r, 4, unit_cost, fmt=CURRENCY_FMT)
        _formula(ws, r, 5, f"=C{r}*D{r}", fmt=CURRENCY_FMT)

    last_item_row = 11 + len(line_items) - 1
    total_row = last_item_row + 2

    _label(ws, total_row, 1, "DETAILED REHAB TOTAL")
    _result(ws, total_row, 5, f"=SUM(E11:E{last_item_row})", fmt=CURRENCY_FMT)

    # --- Which to use ---
    _section_row(ws, total_row + 2, "Rehab Estimate to Use", cols=5)
    _label(ws, total_row + 3, 1, "Use detailed if available?")
    _input(ws, total_row + 3, 2, "Yes")

    _label(ws, total_row + 4, 1, "Selected rehab estimate")
    _result(
        ws, total_row + 4, 2,
        f'=IF(B{total_row + 3}="Yes",E{total_row},B7)',
        fmt=CURRENCY_FMT,
    )

    ws.sheet_properties.tabColor = ACCENT_GREEN
    return ws


def build_deal_calc(wb):
    """DEAL_CALC – main underwriting sheet tying everything together."""
    ws = wb.create_sheet("DEAL_CALC")
    _set_col_widths(ws, [32, 20, 6, 32, 20])

    _header_row(ws, 1, "DEAL CALCULATOR – Underwriting", cols=5)

    # --- Property info ---
    _section_row(ws, 3, "Property Information", cols=5)
    prop_fields = [
        ("Property address", ""),
        ("City / Zip", "Dallas, TX"),
        ("County", "Dallas"),
        ("Sq ft", 1400),
        ("Beds / Baths", "3 / 2"),
        ("Year built", 1965),
        ("Lot size (sqft)", 6500),
        ("Seller ask price", 85000),
    ]
    for i, (lbl, val) in enumerate(prop_fields):
        r = 4 + i
        _label(ws, r, 1, lbl)
        fmt = CURRENCY_FMT if "price" in lbl.lower() else None
        _input(ws, r, 2, val, fmt=fmt)

    # --- ARV ---
    _section_row(ws, 13, "After-Repair Value (ARV)", cols=5)
    _label(ws, 14, 1, "Comp-based ARV (from COMPS)")
    _formula(ws, 14, 2, "=COMPS!B17", fmt=CURRENCY_FMT)

    _label(ws, 15, 1, "Manual ARV override")
    _input(ws, 15, 2, 180000, fmt=CURRENCY_FMT)

    _label(ws, 16, 1, "ARV to use")
    _result(
        ws, 16, 2,
        "=IF(B15>0,B15,B14)",
        fmt=CURRENCY_FMT,
    )

    # --- Repairs ---
    _section_row(ws, 18, "Repair Estimate", cols=5)
    _label(ws, 19, 1, "Rehab estimate (from REPAIRS)")
    _formula(ws, 19, 2, "=REPAIRS!B30", fmt=CURRENCY_FMT)

    # --- Classic MAO ---
    _section_row(ws, 21, "Classic MAO / ROS Ceiling", cols=5)
    _label(ws, 22, 1, "MAO % of ARV (from SETTINGS)")
    _formula(ws, 22, 2, "=SETTINGS!B10", fmt=PCT_FMT)

    _label(ws, 23, 1, "Classic MAO ceiling")
    _result(ws, 23, 2, "=(B16*B22)-B19", fmt=CURRENCY_FMT)

    # --- Precision investor max ---
    _section_row(ws, 25, "Precision Investor-Max Ceiling", cols=5)

    _label(ws, 26, 1, "Realtor commission")
    _formula(ws, 26, 2, "=B16*SETTINGS!B4", fmt=CURRENCY_FMT)

    _label(ws, 27, 1, "Seller closing costs")
    _formula(ws, 27, 2, "=B16*SETTINGS!B5", fmt=CURRENCY_FMT)

    _label(ws, 28, 1, "Hold costs (total)")
    _formula(ws, 28, 2, "=B16*SETTINGS!B6*SETTINGS!B7", fmt=CURRENCY_FMT)

    _label(ws, 29, 1, "Repairs")
    _formula(ws, 29, 2, "=B19", fmt=CURRENCY_FMT)

    _label(ws, 30, 1, "Target buyer profit")
    _formula(ws, 30, 2, "=B16*SETTINGS!B12", fmt=CURRENCY_FMT)

    _label(ws, 31, 1, "Total investor costs")
    _formula(ws, 31, 2, "=SUM(B26:B30)", fmt=CURRENCY_FMT)

    _label(ws, 32, 1, "Precision investor max")
    _result(ws, 32, 2, "=B16-B31", fmt=CURRENCY_FMT)

    # --- Recommended ceiling ---
    _section_row(ws, 34, "Recommended Contract Ceiling", cols=5)
    _label(ws, 35, 1, "Classic MAO ceiling")
    _formula(ws, 35, 2, "=B23", fmt=CURRENCY_FMT)

    _label(ws, 36, 1, "Precision investor max")
    _formula(ws, 36, 2, "=B32", fmt=CURRENCY_FMT)

    _label(ws, 37, 1, "CONTRACT CEILING (lower of two)")
    _result(ws, 37, 2, "=MIN(B35,B36)", fmt=CURRENCY_FMT)

    # --- Offer strategy ---
    _section_row(ws, 39, "Offer Strategy", cols=5)

    _label(ws, 40, 1, "Opening offer")
    _result(
        ws, 40, 2,
        "=B37*(1-SETTINGS!B14)",
        fmt=CURRENCY_FMT,
    )

    _label(ws, 41, 1, "Target offer")
    _result(
        ws, 41, 2,
        "=B37*(1-SETTINGS!B15)",
        fmt=CURRENCY_FMT,
    )

    _label(ws, 42, 1, "Absolute max offer")
    _result(
        ws, 42, 2,
        "=B37*(1-SETTINGS!B16)",
        fmt=CURRENCY_FMT,
    )

    # --- Assignment analysis ---
    _section_row(ws, 44, "Assignment / Wholesale Spread", cols=5)

    _label(ws, 45, 1, "Contract price (your offer)")
    _input(ws, 45, 2, fmt=CURRENCY_FMT)

    _label(ws, 46, 1, "Investor buyer price")
    _formula(ws, 46, 2, "=B32", fmt=CURRENCY_FMT)

    _label(ws, 47, 1, "Assignment fee")
    _result(ws, 47, 2, "=B46-B45", fmt=CURRENCY_FMT)

    _label(ws, 48, 1, "Min assignment fee (SETTINGS)")
    _formula(ws, 48, 2, "=SETTINGS!B13", fmt=CURRENCY_FMT)

    _label(ws, 49, 1, "Fee meets minimum?")
    _formula(
        ws, 49, 2,
        '=IF(B47>=B48,"YES","NO — below min")',
    )

    _label(ws, 50, 1, "Investor ROI")
    _formula(
        ws, 50, 2,
        "=IF(B46>0,(B16-B31-B46)/B46,0)",
        fmt=PCT_FMT,
    )

    _label(ws, 51, 1, "ROI check")
    _formula(
        ws, 51, 2,
        '=IF(B50>=SETTINGS!B12,"PASS","FAIL — ROI too low")',
    )

    ws.sheet_properties.tabColor = ACCENT_GREEN
    return ws


def build_dispo(wb):
    """DISPO_SUMMARY – one-page buyer-facing snapshot."""
    ws = wb.create_sheet("DISPO_SUMMARY")
    _set_col_widths(ws, [30, 22, 6, 30, 22])

    _header_row(ws, 1, "BUYER DEAL SUMMARY", cols=5)

    # Left column: property + numbers
    _section_row(ws, 3, "Property Details", cols=2)
    buyer_fields_left = [
        ("Address", "=DEAL_CALC!B4"),
        ("City / Zip", "=DEAL_CALC!B5"),
        ("Sq ft", "=DEAL_CALC!B7"),
        ("Beds / Baths", "=DEAL_CALC!B8"),
        ("Year built", "=DEAL_CALC!B9"),
    ]
    for i, (lbl, val) in enumerate(buyer_fields_left):
        r = 4 + i
        _label(ws, r, 1, lbl)
        _formula(ws, r, 2, val)

    _section_row(ws, 10, "Investment Numbers", cols=2)
    buyer_numbers = [
        ("ARV", "=DEAL_CALC!B16", CURRENCY_FMT),
        ("Est. rehab", "=DEAL_CALC!B19", CURRENCY_FMT),
        ("Purchase price", "=DEAL_CALC!B46", CURRENCY_FMT),
        ("Total investment", "=DEAL_CALC!B46+DEAL_CALC!B19", CURRENCY_FMT),
        ("Projected profit", "=DEAL_CALC!B16-DEAL_CALC!B31-DEAL_CALC!B46", CURRENCY_FMT),
        ("Est. ROI", "=DEAL_CALC!B50", PCT_FMT),
    ]
    for i, (lbl, val, fmt) in enumerate(buyer_numbers):
        r = 11 + i
        _label(ws, r, 1, lbl)
        _result(ws, r, 2, val, fmt=fmt)

    # Right column: notes
    ws.merge_cells("D3:E3")
    notes_cell = ws.cell(row=3, column=4, value="Notes for Buyer")
    notes_cell.font = SECTION_FONT
    notes_cell.fill = SECTION_FILL
    notes_cell.alignment = Alignment(horizontal="left", vertical="center")

    _label(ws, 4, 4, "Comparable sales")
    _input(ws, 4, 5, "See COMPS tab")
    _label(ws, 5, 4, "Scope of work")
    _input(ws, 5, 5, "See REPAIRS tab")
    _label(ws, 6, 4, "Title status")
    _input(ws, 6, 5, "Clear — see title report")
    _label(ws, 7, 4, "Assignment / close date")
    _input(ws, 7, 5, "TBD")
    _label(ws, 8, 4, "EMD required")
    _input(ws, 8, 5, "$2,000")

    # Disclaimer
    disc_row = 18
    ws.merge_cells(
        start_row=disc_row, start_column=1, end_row=disc_row + 2, end_column=5
    )
    disc = ws.cell(
        row=disc_row,
        column=1,
        value=(
            "DISCLAIMER: This summary is for informational purposes only "
            "and does not constitute legal, tax, or investment advice. "
            "Buyer should conduct independent due diligence. All figures "
            "are estimates and may vary. Consult licensed professionals "
            "before closing."
        ),
    )
    disc.font = Font(name="Calibri", size=9, italic=True, color="666666")
    disc.alignment = Alignment(wrap_text=True, vertical="top")

    ws.sheet_properties.tabColor = "FFC000"
    return ws


# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------

def generate_workbook(output_path=None):
    """Build and save the deal calculator workbook. Returns the file path."""
    if output_path is None:
        output_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "Deal_Calculator_Workbook.xlsx",
        )

    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    # Build sheets in tab order
    build_settings(wb)
    build_comps(wb)
    build_repairs(wb)
    build_deal_calc(wb)
    build_dispo(wb)

    # Set DEAL_CALC as the active sheet on open
    wb.active = wb.sheetnames.index("DEAL_CALC")

    wb.save(output_path)
    print(f"Workbook saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    generate_workbook()
